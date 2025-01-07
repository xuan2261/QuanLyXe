import streamlit as st
from modules.vehicle import manage_vehicles
from modules.booking import manage_bookings
from config import db
import pandas as pd
import matplotlib.pyplot as plt
from bson import ObjectId
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

def admin_dashboard(user):
    st.subheader("Quản lý hệ thống: Admin")
    menu = ["Quản Lý Xe", "Quản Lý Đơn Đặt Hàng", "Thống Kê"]
    # Lưu lựa chọn vào session_state
    if 'selected_menu' not in st.session_state:
        st.session_state['selected_menu'] = menu[0]

    # Thêm key="menu_admin" cho st.sidebar.selectbox
    choice = st.sidebar.selectbox("Menu Quản Lý", menu, index=menu.index(st.session_state['selected_menu']), key="menu_admin")
    st.session_state['selected_menu'] = choice

    if choice == "Quản Lý Xe":
        manage_vehicles()  # Hàm quản lý xe
    elif choice == "Quản Lý Đơn Đặt Hàng":
        manage_bookings()  # Gọi hàm quản lý đơn đặt hàng
    elif choice == "Thống Kê":
        view_statistics()  # Gọi hàm thống kê

def view_statistics():
    st.subheader("Thống Kê Chi Tiết")

    # Bộ lọc
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Ngày bắt đầu", datetime.date.today() - datetime.timedelta(days=30))
    with col2:
        end_date = st.date_input("Ngày kết thúc", datetime.date.today())

    # Chuyển đổi kiểu dữ liệu
    start_date = datetime.datetime.combine(start_date, datetime.datetime.min.time())
    end_date = datetime.datetime.combine(end_date, datetime.datetime.max.time())

    # Thêm bộ lọc xe
    vehicle_options = [f"{v['brand']} {v['model']} - {v['license_plate']}" for v in db.vehicles.find({}, {"brand": 1, "model": 1, "license_plate": 1})]
    selected_vehicles = st.multiselect("Chọn xe", vehicle_options)

    # Tạo bộ lọc chung cho các truy vấn
    query_filter = {
        "created_at": {"$gte": start_date, "$lte": end_date}
    }

    if selected_vehicles:
        # Lấy danh sách biển số xe từ lựa chọn của người dùng
        selected_license_plates = [v.split(" - ")[-1] for v in selected_vehicles]
        vehicle_ids = [v["_id"] for v in db.vehicles.find({"license_plate": {"$in": selected_license_plates}}, {"_id": 1})]
        query_filter["vehicle_id"] = {"$in": vehicle_ids}

    # Thống kê tổng số đơn đặt hàng
    total_bookings = db.bookings.count_documents(query_filter)
    st.metric(label="Tổng số đơn đặt hàng", value=total_bookings)

    # Thống kê tổng doanh thu
    total_revenue = db.bookings.aggregate([
        {"$match": query_filter},
        {"$group": {"_id": None, "total": {"$sum": "$total_price"}}}
    ])
    revenue = next(total_revenue, {}).get("total", 0)
    st.metric(label="Tổng doanh thu (USD)", value=f"{revenue:,.2f}") # Format số

    # Thống kê số lượng khách hàng
    # Lưu ý: Cần lọc các đơn hàng của cùng 1 khách hàng, nếu không sẽ tính trùng
    distinct_customers = db.bookings.aggregate([
        {"$match": query_filter},
        {"$group": {"_id": "$user_id"}}
    ])
    total_customers = len(list(distinct_customers))
    st.metric(label="Tổng số khách hàng", value=total_customers)

    # Thống kê tổng số xe
    total_vehicles = db.vehicles.count_documents({})
    st.metric(label="Tổng số xe", value=total_vehicles)

    # Thống kê doanh thu theo tháng
    st.subheader("Doanh Thu Theo Tháng")
    monthly_revenue = db.bookings.aggregate([
        {"$match": query_filter},
        {"$group": {
            "_id": {
                "year": {"$year": "$created_at"},
                "month": {"$month": "$created_at"}
            },
            "total": {"$sum": "$total_price"}
        }},
        {"$sort": {"_id.year": 1, "_id.month": 1}}
    ])
    monthly_revenue = list(monthly_revenue)

    # Tạo DataFrame từ dữ liệu truy vấn
    revenue_data = []
    for item in monthly_revenue:
        month_str = f"{item['_id']['year']}-{item['_id']['month']:02}"
        revenue_data.append({"Tháng": month_str, "Doanh Thu (USD)": item["total"]})
    revenue_df = pd.DataFrame(revenue_data)

    # Kiểm tra nếu chỉ có một điểm dữ liệu
    if len(revenue_df) == 1:
        # Lấy tháng hiện tại
        current_month = datetime.datetime.strptime(revenue_df.iloc[0]["Tháng"], "%Y-%m")
        previous_month = (current_month - datetime.timedelta(days=30)).strftime("%Y-%m")
        revenue_df = pd.concat([pd.DataFrame([{"Tháng": previous_month, "Doanh Thu (USD)": 0}]), revenue_df], ignore_index=True)

    # Hiển thị biểu đồ
    if not revenue_df.empty:
        revenue_df = revenue_df.sort_values(by="Tháng")
        st.line_chart(revenue_df.set_index("Tháng"))
    else:
        st.write("Không có dữ liệu doanh thu trong khoảng thời gian này.")

    # Thống kê số lượng đơn đặt hàng theo trạng thái thanh toán
    st.subheader("Số Lượng Đơn Đặt Hàng Theo Trạng Thái Thanh Toán")
    booking_status = db.bookings.aggregate([
        {"$match": query_filter},
        {"$group": {"_id": "$payment_status", "count": {"$sum": 1}}}
    ])
    booking_status_list = list(booking_status)  # Chuyển đổi cursor thành list
    statuses = [item["_id"] for item in booking_status_list]
    counts = [item["count"] for item in booking_status_list]
    status_df = pd.DataFrame({"Trạng Thái": statuses, "Số Lượng": counts})
    st.bar_chart(status_df.set_index("Trạng Thái"))

    # Thống kê doanh thu theo xe
    st.subheader("Doanh Thu Theo Từng Xe")
    vehicle_revenue = db.bookings.aggregate([
        {"$match": query_filter},
        {"$lookup": {
            "from": "vehicles",
            "localField": "vehicle_id",
            "foreignField": "_id",
            "as": "vehicle_info"
        }},
        {"$unwind": "$vehicle_info"},
        {"$group": {
            "_id": "$vehicle_info.license_plate",
            "total_revenue": {"$sum": "$total_price"}
        }},
        {"$sort": {"total_revenue": -1}}
    ])
    vehicle_revenue = list(vehicle_revenue)
    vehicles = [item["_id"] for item in vehicle_revenue]
    revenues_per_vehicle = [item["total_revenue"] for item in vehicle_revenue]

    # Tạo DataFrame
    vehicle_df = pd.DataFrame({"Biển Số": vehicles, "Doanh Thu (USD)": revenues_per_vehicle})
    # Sắp xếp theo doanh thu
    vehicle_df = vehicle_df.sort_values(by="Doanh Thu (USD)", ascending=False)
    # Vẽ biểu đồ
    st.bar_chart(vehicle_df.set_index("Biển Số"))


    # Thống kê số lượng đơn đặt hàng theo tháng
    st.subheader("Số Lượng Đơn Đặt Hàng Theo Tháng")
    monthly_booking_count = db.bookings.aggregate([
        {"$match": query_filter},
        {"$addFields": {"month": {"$month": "$created_at"}}},  # Trích xuất tháng
        {"$group": {
            "_id": {
                "year": {"$year": "$created_at"},
                "month": {"$month": "$created_at"}
            },
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id.year": 1, "_id.month": 1}}
    ])
    monthly_booking_count = list(monthly_booking_count)
    # print("monthly_booking_count:", monthly_booking_count)
    # Tạo DataFrame từ dữ liệu truy vấn
    booking_count_data = []
    for item in monthly_booking_count:
        month_str = f"{item['_id']['year']}-{item['_id']['month']:02}"  # Định dạng tháng
        booking_count_data.append({"Tháng": month_str, "Số Lượng Đơn": item["count"]})
    booking_count_df = pd.DataFrame(booking_count_data)

    # Thêm dữ liệu giả nếu chỉ có một tháng
    if len(booking_count_df) == 1:
        # Lấy tháng hiện tại
        current_month = datetime.datetime.strptime(booking_count_df['Tháng'][0], "%Y-%m")
        # Thêm tháng trước đó với giá trị 0
        previous_month = current_month - datetime.timedelta(days=1)
        previous_month_str = previous_month.strftime("%Y-%m")
        booking_count_df = pd.concat([pd.DataFrame([{"Tháng": previous_month_str, "Số Lượng Đơn": 0}]), booking_count_df], ignore_index=True)

    # Kiểm tra xem DataFrame có trống không
    if not booking_count_df.empty:
        st.line_chart(booking_count_df.set_index("Tháng"))
    else:
        st.write("Không có dữ liệu đơn đặt hàng trong khoảng thời gian này.")
    
    # Thống kê tỷ lệ phần trăm trạng thái thanh toán
    st.subheader("Tỷ Lệ Phần Trăm Trạng Thái Thanh Toán")
    payment_status_stats = db.bookings.aggregate([
        {"$match": query_filter},
        {"$group": {"_id": "$payment_status", "count": {"$sum": 1}}},
        {"$project": {
            "payment_status": "$_id",
            "count": 1,
            "percentage": {
                "$cond": {
                    "if": {"$eq": [total_bookings, 0]},
                    "then": 0,
                    "else": {"$multiply": [{"$divide": ["$count", total_bookings]}, 100]}
                }
            }
        }}
    ])

    payment_status_stats_df = pd.DataFrame(list(payment_status_stats))
    # Kiểm tra xem cột 'percentage' có tồn tại không trước khi format
    if 'percentage' in payment_status_stats_df.columns:
        payment_status_stats_df["percentage"] = payment_status_stats_df["percentage"].map("{:.2f}%".format)
    st.dataframe(payment_status_stats_df)
    
    # Thống kê tình trạng xe
    st.subheader("Tình Trạng Xe")
    vehicle_statuses = db.vehicles.aggregate([
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ])
    vehicle_statuses = list(vehicle_statuses)
    statuses = [item["_id"] for item in vehicle_statuses]
    counts = [item["count"] for item in vehicle_statuses]
    status_vehicle_df = pd.DataFrame({"Trạng Thái": statuses, "Số Lượng": counts})
    
    # Vẽ biểu đồ cột
    st.bar_chart(status_vehicle_df.set_index("Trạng Thái"))

    # Xuất báo cáo
    if st.button("Xuất Báo Cáo (Excel)"):
        output = export_to_excel(
            revenue_df,
            status_df,
            vehicle_df,
            booking_count_df,
            payment_status_stats_df,
            status_vehicle_df
        )
        st.download_button(
            label="Tải xuống",
            data=output,
            file_name="thong_ke.xlsx",
            mime="application/vnd.ms-excel"
        )

def export_to_excel(revenue_df, status_df, vehicle_df, booking_count_df, payment_status_stats_df, status_vehicle_df):
    output = BytesIO()
    workbook = Workbook()

    # Định dạng chung
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    # Sheet Doanh Thu Theo Tháng
    sheet_revenue = workbook.active
    sheet_revenue.title = "Doanh Thu Theo Tháng"
    for r in dataframe_to_rows(revenue_df, index=True, header=True):
        sheet_revenue.append(r)
    # Format header
    for cell in sheet_revenue["1:1"]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border
    # Căn giữa cột index
    for row in sheet_revenue.iter_rows():
        row[0].alignment = center_alignment
        for cell in row:
            cell.border = border

    # Sheet Số Lượng Đơn Đặt Hàng Theo Trạng Thái Thanh Toán
    sheet_status = workbook.create_sheet("SL Đơn Hàng Theo Trạng Thái")
    for r in dataframe_to_rows(status_df, index=True, header=True):
        sheet_status.append(r)
    for cell in sheet_status["1:1"]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border
    for row in sheet_status.iter_rows():
        row[0].alignment = center_alignment
        for cell in row:
            cell.border = border

    # Sheet Doanh Thu Theo Từng Xe
    sheet_vehicle = workbook.create_sheet("Doanh Thu Theo Từng Xe")
    for r in dataframe_to_rows(vehicle_df, index=True, header=True):
        sheet_vehicle.append(r)
    for cell in sheet_vehicle["1:1"]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border
    for row in sheet_vehicle.iter_rows():
        row[0].alignment = center_alignment
        for cell in row:
            cell.border = border

    # Sheet Số Lượng Đơn Đặt Hàng Theo Tháng
    sheet_booking_count = workbook.create_sheet("SL Đơn Hàng Theo Tháng")
    for r in dataframe_to_rows(booking_count_df, index=True, header=True):
        sheet_booking_count.append(r)
    for cell in sheet_booking_count["1:1"]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border
    for row in sheet_booking_count.iter_rows():
        row[0].alignment = center_alignment
        for cell in row:
            cell.border = border

    # Sheet Tỷ Lệ Phần Trăm Trạng Thái Thanh Toán
    sheet_payment_status = workbook.create_sheet("Tỷ Lệ Trạng Thái TT")
    for r in dataframe_to_rows(payment_status_stats_df, index=False, header=True):
        sheet_payment_status.append(r)
    for cell in sheet_payment_status["1:1"]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border
    for row in sheet_payment_status.iter_rows():
        for cell in row:
            cell.border = border
            
    # Sheet Tình Trạng Xe
    sheet_vehicle_status = workbook.create_sheet("Tình Trạng Xe")
    for r in dataframe_to_rows(status_vehicle_df, index=True, header=True):
        sheet_vehicle_status.append(r)
    for cell in sheet_vehicle_status["1:1"]:
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border
    for row in sheet_vehicle_status.iter_rows():
        row[0].alignment = center_alignment
        for cell in row:
            cell.border = border

    workbook.save(output)
    return output